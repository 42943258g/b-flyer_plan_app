import os
import re
import json
import tempfile
import shutil
import base64
import hashlib
import secrets
from pathlib import Path
from typing import Any, Dict, List, Optional
from copy import copy
from collections import defaultdict

import psycopg
from psycopg.types.json import Jsonb

from fastapi import FastAPI, HTTPException, BackgroundTasks, Body, Request
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from starlette.middleware.sessions import SessionMiddleware

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
import uuid
import unicodedata

from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill

from openpyxl.worksheet.cell_range import CellRange

import io
import time

# --- Formula evaluation for preview (formulas) ---
# xlcalculator は古く(最終 2023)、Python 3.13 だと入らない/動かないことがあるため、
# formulas を優先して使う（Excel計算エンジン）
try:
    import formulas  # pip install formulas[excel] もしくは formulas[all]
    _HAS_FORMULAS = True
except Exception:
    formulas = None
    _HAS_FORMULAS = False


import datetime as _dt
from fastapi import UploadFile, File
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell.cell import MergedCell

from openpyxl.utils.cell import range_boundaries

from datetime import timezone, timedelta
JST = timezone(timedelta(hours=9))

# ===== Dev / Debug =====
# ログインを完全にバイパスする開発用モード。
# 事故防止のため「DEV_NO_LOGIN=1」が必須。
# Render 本番で誤って有効化しないよう、Render 環境では
# さらに「ALLOW_DEV_NO_LOGIN_ON_RENDER=1」が無い限り無効化。
def dev_no_login_enabled() -> bool:
    # NOTE: parse_bool はこの後に定義されるが、Python は関数実行時に解決するのでOK。
    if not parse_bool(os.environ.get("DEV_NO_LOGIN")):
        return False
    # Render では明示許可が無い限りOFF
    if os.environ.get("RENDER") and not parse_bool(os.environ.get("ALLOW_DEV_NO_LOGIN_ON_RENDER")):
        return False
    return True


def _dev_user() -> dict:
    # dev_no_login 時は常に admin 扱い（開発しやすさ優先）
    return {"username": "dev", "role": "admin"}


APP_DIR = Path(__file__).resolve().parent
INDEX_HTML_PATH = APP_DIR / "index.html"
XLSX_PATH = APP_DIR / "list_format.xlsx"

DB_URL_TXT = APP_DIR / "database_url.txt"
DOTENV_PATH = APP_DIR / ".env"
SECRET_TXT = APP_DIR / "secret_key.txt"

_FULLWIDTH_TRANS = str.maketrans("０１２３４５６７８９．－", "0123456789.-")

# ===== 追加テーブル名 =====
LOGIN_USERS_TARGET = "login_users"
UPDATE_LOG_TARGET = "table_update_log"

ALLOWED_TARGETS = ("shop_master", "syoken", "schedule", LOGIN_USERS_TARGET, UPDATE_LOG_TARGET)

ROLE_ORDER = {"viewer": 0, "editor": 1, "admin": 2}


def _read_text(path: Path) -> Optional[str]:
    if not path.is_file():
        return None
    s = path.read_text(encoding="utf-8").strip()
    return s or None


def _read_database_url_from_dotenv() -> Optional[str]:
    if not DOTENV_PATH.is_file():
        return None
    for line in DOTENV_PATH.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" not in line:
            continue
        k, v = line.split("=", 1)
        if k.strip() == "DATABASE_URL":
            v = v.strip().strip('"').strip("'")
            return v or None
    return None


def get_database_url() -> Optional[str]:
    return (
        os.environ.get("DATABASE_URL")
        or _read_text(DB_URL_TXT)
        or _read_database_url_from_dotenv()
    )


def db_url_or_die() -> str:
    url = get_database_url()
    if not url:
        raise RuntimeError(
            "DATABASE_URL が見つかりません。\n"
            "いずれかで設定してください:\n"
            "  (1) 環境変数 DATABASE_URL\n"
            "  (2) app.py と同フォルダに database_url.txt（1行でURL）\n"
            "  (3) app.py と同フォルダに .env（DATABASE_URL=...）\n"
        )
    return url


def get_secret_key() -> str:
    # 本番は環境変数 SECRET_KEY 推奨。無ければ secret_key.txt、最後に自動生成。
    s = os.environ.get("SECRET_KEY") or _read_text(SECRET_TXT)
    if s:
        return s
    # 自動生成（再起動でセッション無効になる）
    return secrets.token_urlsafe(32)


app = FastAPI()
app.add_middleware(
    SessionMiddleware,
    secret_key=get_secret_key(),
    same_site="lax",
    https_only=False,  # Renderでhttps運用なら True でもOK
)


# ===== Preview workbook cache (for live formula recalculation in preview) =====
# preview_id -> {"bytes": xlsx_bytes, "sheet": str, "formula_cells": list[tuple[int,str]], "updated_at": float}
_PREVIEW_CACHE: dict[str, dict] = {}
_PREVIEW_CACHE_MAX = 20
_PREVIEW_CACHE_TTL_SEC = 60 * 60  # 1 hour


def _preview_cache_prune():
    now = time.time()
    # TTL prune
    dead = [k for k, v in _PREVIEW_CACHE.items() if now - float(v.get("updated_at", 0)) > _PREVIEW_CACHE_TTL_SEC]
    for k in dead:
        _PREVIEW_CACHE.pop(k, None)
    # size prune
    if len(_PREVIEW_CACHE) > _PREVIEW_CACHE_MAX:
        items = sorted(_PREVIEW_CACHE.items(), key=lambda kv: float(kv[1].get("updated_at", 0)))
        for k, _ in items[: max(0, len(items) - _PREVIEW_CACHE_MAX)]:
            _PREVIEW_CACHE.pop(k, None)


def _first_visible_sheet(wb):
    for s in wb.worksheets:
        if getattr(s, "sheet_state", "visible") == "visible":
            return s
    return wb.active


def _extract_formula_cells_from_bytes(content: bytes) -> tuple[str, list[tuple[int, str]]]:
    # どのセルが数式か（=...）を可視シートから拾う
    wb = load_workbook(io.BytesIO(content), data_only=False)
    ws = _first_visible_sheet(wb)

    MAX_COL = column_index_from_string("BY")
    MAX_ROW = 1200
    max_row = min(ws.max_row or 1, MAX_ROW)
    max_col = min(ws.max_column or 1, MAX_COL)

    out: list[tuple[int, str]] = []
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            # openpyxl は数式セルだと data_type == "f"
            if getattr(cell, "data_type", None) == "f":
                out.append((r, get_column_letter(c)))
            else:
                v = getattr(cell, "value", None)
                if isinstance(v, str) and v.startswith("="):
                    out.append((r, get_column_letter(c)))
    return ws.title, out


def _preview_cache_put(content: bytes) -> tuple[str, dict]:
    _preview_cache_prune()
    pid = uuid.uuid4().hex
    if ENABLE_PREVIEW_RECALC:
        sheet, formula_cells = _extract_formula_cells_from_bytes(content)
    else:
        sheet, formula_cells = "", []
    meta = {"bytes": content, "sheet": sheet, "formula_cells": formula_cells, "updated_at": time.time()}
    _PREVIEW_CACHE[pid] = meta
    return pid, meta


def _preview_cache_get(pid: str) -> dict:
    _preview_cache_prune()
    meta = _PREVIEW_CACHE.get(pid)
    if not meta:
        raise HTTPException(404, "プレビューが期限切れです。もう一度「新規作成」または「Excel取込」をしてください。")
    meta["updated_at"] = time.time()
    return meta


def _as_number_or_text(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v

    s = str(v).strip()
    if s == "":
        return None

    s2 = s.translate(_FULLWIDTH_TRANS).replace(",", "")

    if re.fullmatch(r"-?\d+", s2):
        try:
            return int(s2)
        except Exception:
            return s
    if re.fullmatch(r"-?\d+\.\d+", s2):
        try:
            return float(s2)
        except Exception:
            return s
    return s


def _is_empty_row(d: dict) -> bool:
    if not isinstance(d, dict) or len(d) == 0:
        return True
    for v in d.values():
        if v is None:
            continue
        if str(v).strip() != "":
            return False
    return True


def parse_bool(v) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in ("1", "true", "t", "yes", "y", "on")

ENABLE_PREVIEW_RECALC = parse_bool(os.environ.get("ENABLE_PREVIEW_RECALC", "0"))
PREVIEW_CALC_ON_LOAD = False


PREVIEW_CALC_ON_LOAD = False  # 計算は重いのでプレビューでは無効

def normalize_role(v: Any) -> str:
    s = ("" if v is None else str(v).strip().lower())
    if s in ROLE_ORDER:
        return s
    return "viewer"


def session_user(request: Request) -> Optional[dict]:
    u = request.session.get("user")
    if isinstance(u, dict):
        return u
    # DEV_NO_LOGIN=1 のときはログイン無しで入れる
    if dev_no_login_enabled():
        return _dev_user()
    return None


def require_login(request: Request) -> dict:
    u = session_user(request)
    if not u:
        raise HTTPException(401, "not logged in")
    return u


def require_role(request: Request, min_role: str) -> dict:
    u = require_login(request)
    role = normalize_role(u.get("role"))
    if ROLE_ORDER.get(role, 0) < ROLE_ORDER.get(min_role, 0):
        raise HTTPException(403, f"forbidden: require role >= {min_role}")
    return u


def hash_password(password: str, iterations: int = 260_000) -> str:
    """
    pbkdf2_sha256$<iterations>$<salt_b64>$<hash_b64>
    """
    pw = (password or "").strip()
    if pw == "":
        raise ValueError("empty password")

    salt = secrets.token_bytes(16)
    dk = hashlib.pbkdf2_hmac("sha256", pw.encode("utf-8"), salt, iterations, dklen=32)

    salt_b64 = base64.urlsafe_b64encode(salt).decode("ascii").rstrip("=")
    hash_b64 = base64.urlsafe_b64encode(dk).decode("ascii").rstrip("=")

    return f"pbkdf2_sha256${iterations}${salt_b64}${hash_b64}"


def verify_password(password: str, stored: str) -> bool:
    """
    stored: pbkdf2_sha256$iters$salt_b64$hash_b64
    """
    if not stored:
        return False
    s = str(stored).strip()
    if not s.startswith("pbkdf2_sha256$"):
        # 旧データ救済（平文が入ってた等）を許さないなら False にしてOK
        return False

    try:
        _, iters, salt_b64, hash_b64 = s.split("$", 3)
        iterations = int(iters)

        # base64 padding復元
        def _b64dec(x: str) -> bytes:
            pad = "=" * ((4 - (len(x) % 4)) % 4)
            return base64.urlsafe_b64decode(x + pad)

        salt = _b64dec(salt_b64)
        expected = _b64dec(hash_b64)

        dk = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iterations, dklen=len(expected))
        return secrets.compare_digest(dk, expected)
    except Exception:
        return False


def connect_db() -> psycopg.Connection:
    url = getattr(app.state, "database_url", None) or db_url_or_die()
    return psycopg.connect(url)


def ensure_tables(conn: psycopg.Connection):
    conn.execute("""
    create table if not exists shop_master (
      row_no integer,
      payload jsonb not null,
      updated_at timestamptz default now()
    );
    """)
    conn.execute("create index if not exists shop_master_row_no_idx on shop_master(row_no);")

    conn.execute("""
    create table if not exists syoken (
      row_no integer,
      payload jsonb not null,
      updated_at timestamptz default now()
    );
    """)
    conn.execute("create index if not exists syoken_row_no_idx on syoken(row_no);")

    conn.execute("""
    create table if not exists schedule (
      row_no integer,
      payload jsonb not null,
      updated_at timestamptz default now()
    );
    """)
    conn.execute("create index if not exists schedule_row_no_idx on schedule(row_no);")

    conn.execute(f"""
    create table if not exists {LOGIN_USERS_TARGET} (
      row_no integer,
      payload jsonb not null,
      updated_at timestamptz default now()
    );
    """)
    conn.execute(f"create index if not exists {LOGIN_USERS_TARGET}_row_no_idx on {LOGIN_USERS_TARGET}(row_no);")

    conn.execute("""
    create table if not exists table_schema (
      target text not null,
      columns jsonb not null,
      updated_at timestamptz default now()
    );
    """)

    conn.execute(f"""
    create table if not exists {UPDATE_LOG_TARGET} (
      id bigserial primary key,
      target text not null,
      saved_rows integer not null,
      columns jsonb not null,
      actor_username text,
      client_ip text,
      user_agent text,
      happened_at timestamptz default now()
    );
    """)
    conn.execute(f"create index if not exists {UPDATE_LOG_TARGET}_happened_at_idx on {UPDATE_LOG_TARGET}(happened_at desc);")
    conn.execute(f"create index if not exists {UPDATE_LOG_TARGET}_target_idx on {UPDATE_LOG_TARGET}(target);")
    # 既存DBに後から列を追加できるように（マイグレーション）
    conn.execute(f"alter table {UPDATE_LOG_TARGET} add column if not exists actor_username text;")
    conn.execute(f"alter table {UPDATE_LOG_TARGET} add column if not exists client_ip text;")
    conn.execute(f"alter table {UPDATE_LOG_TARGET} add column if not exists user_agent text;")
    conn.execute(f"alter table {UPDATE_LOG_TARGET} add column if not exists columns jsonb;")
    conn.execute(f"alter table {UPDATE_LOG_TARGET} add column if not exists happened_at timestamptz default now();")
    conn.execute(f"alter table {UPDATE_LOG_TARGET} add column if not exists diff jsonb;")
    conn.execute(f"alter table {UPDATE_LOG_TARGET} add column if not exists added_count integer;")
    conn.execute(f"alter table {UPDATE_LOG_TARGET} add column if not exists deleted_count integer;")
    conn.execute(f"alter table {UPDATE_LOG_TARGET} add column if not exists batch_id uuid;")
    conn.execute(f"alter table {UPDATE_LOG_TARGET} add column if not exists op text;")          # add / delete
    conn.execute(f"alter table {UPDATE_LOG_TARGET} add column if not exists record jsonb;")     # 変更対象の1レコード


    conn.commit()


@app.on_event("startup")
def startup():
    url = db_url_or_die()
    app.state.database_url = url
    with psycopg.connect(url) as conn:
        ensure_tables(conn)

@app.get("/api/dv-maps")
def api_get_dv_maps(request: Request):
    require_role(request, "viewer")

    # export_xlsx と同じ正規化
    def _norm_shop_code(x: Any) -> str:
        s = ("" if x is None else str(x)).strip()
        s = s.translate(str.maketrans("０１２３４５６７８９", "0123456789"))
        return s

    def _norm_busu_key(x: Any) -> str:
        s = ("" if x is None else str(x)).strip()
        s = s.translate(str.maketrans("０１２３４５６７８９，", "0123456789,"))  # 全角→半角
        s = s.replace(",", "")  # 1,000 → 1000
        n = _as_number_or_text(s)
        if isinstance(n, float) and n.is_integer():
            n = int(n)
        return str(n) if isinstance(n, (int, float)) else s

    def _payload_get(payload: dict, keys: list[str]):
        for k in keys:
            if k in payload:
                return payload.get(k)
        return None

    def _payload_get_any(payload: dict, keys: list[str]):
        for k in keys:
            if k in payload:
                return payload.get(k)
        return None

    SYOKEN_SHOP_CODE_KEYS = ["店番", "店番フィールド", "shop_code", "shop_cd", "店舗コード", "店コード", "店舗番号"]
    SYOKEN_BUSU_KEYS      = ["部数", "部数フィールド", "busu", "BUSU", "枚数"]
    SYOKEN_PATTERN_KEYS   = ["パターン", "パターンフィールド", "pattern", "PATTERN"]

    from collections import defaultdict
    busu_map = defaultdict(list)       # shop -> [busu...]
    pattern_map = defaultdict(list)    # "shop|busu" -> [pattern...]

    with connect_db() as conn:
        rows = conn.execute("select payload from syoken order by row_no").fetchall()

    for (payload,) in rows:
        payload = payload or {}
        sc_raw = _payload_get(payload, SYOKEN_SHOP_CODE_KEYS)
        bs_raw = _payload_get_any(payload, SYOKEN_BUSU_KEYS)
        pt_raw = _payload_get_any(payload, SYOKEN_PATTERN_KEYS)

        sc = _norm_shop_code(sc_raw)
        if not sc or bs_raw is None:
            continue

        bs_key = _norm_busu_key(bs_raw)
        if not bs_key:
            continue

        if bs_key not in busu_map[sc]:
            busu_map[sc].append(bs_key)

        if pt_raw is not None:
            pt = str(pt_raw).strip()
            if pt:
                key = f"{sc}|{bs_key}"
                if pt not in pattern_map[key]:
                    pattern_map[key].append(pt)

    # Excelと同じ列（固定）
    def _cols(start: str, end: str, step: int):
        s = column_index_from_string(start)
        e = column_index_from_string(end)
        out = []
        for c in range(s, e + 1, step):
            out.append(get_column_letter(c))
        return out

    busu_cols = _cols("H", "BP", 4)   # H, L, P ... BP
    pat_cols  = _cols("I", "BQ", 4)   # I, M, Q ... BQ

    return {
        "busu_map": dict(busu_map),
        "pattern_map": dict(pattern_map),
        "busu_cols": busu_cols,
        "pattern_cols": pat_cols,
        "warn_fill": "#FFEB9C",
    }



@app.get("/", response_class=HTMLResponse)
def root():
    if not INDEX_HTML_PATH.is_file():
        raise HTTPException(404, f"index.html not found next to app.py: {INDEX_HTML_PATH}")
    return FileResponse(str(INDEX_HTML_PATH), media_type="text/html; charset=utf-8")


def _columns_from_schema(conn: psycopg.Connection, target: str) -> List[str]:
    sch = conn.execute(
        "select columns from table_schema where target=%s order by updated_at desc limit 1",
        (target,),
    ).fetchone()
    columns = sch[0] if sch else []
    if isinstance(columns, str):
        try:
            columns = json.loads(columns)
        except Exception:
            columns = []
    if not isinstance(columns, list):
        columns = []
    columns = [c for c in columns if isinstance(c, str)]

    if target == LOGIN_USERS_TARGET and not columns:
        return ["username", "password_hash", "role", "is_active"]

    return columns


def _pick_client_ip(request: Request) -> Optional[str]:
    xff = request.headers.get("x-forwarded-for")
    if xff:
        return xff.split(",")[0].strip() or None
    if request.client:
        return request.client.host
    return None


# ===== Auth API =====

@app.get("/api/me")
def me(request: Request):
    u = session_user(request)
    if not u:
        raise HTTPException(401, "not logged in")
    return {"username": u.get("username"), "role": normalize_role(u.get("role"))}


@app.post("/api/logout")
def logout(request: Request):
    request.session.clear()
    return {"ok": True}


@app.post("/api/login")
def login(request: Request, body: Dict[str, Any] = Body(...)):
    username = (body.get("username") or "").strip()
    password = (body.get("password") or "")
    if username == "" or password == "":
        raise HTTPException(400, "username/password required")

    with connect_db() as conn:
        ensure_tables(conn)

        row = conn.execute(
            f"""
            select payload
            from {LOGIN_USERS_TARGET}
            where payload->>'username' = %s
            order by row_no nulls last
            limit 1
            """,
            (username,),
        ).fetchone()

        if not row:
            raise HTTPException(401, "invalid username/password")

        payload = row[0] or {}
        if not isinstance(payload, dict):
            raise HTTPException(401, "invalid username/password")

        if not parse_bool(payload.get("is_active", True)):
            raise HTTPException(403, "account disabled")

        stored_hash = payload.get("password_hash")
        if not verify_password(password, "" if stored_hash is None else str(stored_hash)):
            raise HTTPException(401, "invalid username/password")

        role = normalize_role(payload.get("role"))

    request.session["user"] = {"username": username, "role": role}
    return {"ok": True, "username": username, "role": role}


def _check_table_permission(request: Request, target: str, write: bool):
    # 未ログインはNG
    u = require_login(request)
    role = normalize_role(u.get("role"))

    # 読み取り
    if not write:
        # 更新ログは admin のみ
        if target == UPDATE_LOG_TARGET:
            require_role(request, "admin")
        # login_users は admin のみ
        elif target == LOGIN_USERS_TARGET:
            require_role(request, "admin")
        else:
            require_role(request, "viewer")
        return

    # 書き込み
    if target in ("schedule", "shop_master", "syoken"):
        require_role(request, "editor")
    elif target == LOGIN_USERS_TARGET:
        require_role(request, "admin")
    else:
        raise HTTPException(400, f"{target} is read-only")


@app.get("/api/table/{target}")
def get_table(target: str, request: Request):
    if target not in ALLOWED_TARGETS:
        raise HTTPException(400, f"target must be one of {ALLOWED_TARGETS}")

    _check_table_permission(request, target, write=False)

    with connect_db() as conn:
        ensure_tables(conn)

        if target == UPDATE_LOG_TARGET:
            rows = conn.execute(
                f"""
                select id, happened_at, target, op, batch_id, actor_username, client_ip, user_agent, record
                from {UPDATE_LOG_TARGET}
                order by happened_at desc, id desc
                limit 5000
                """
            ).fetchall()

            out = []
            for r in rows:
                rec = r[8] or {}
                out.append({
                    "id": r[0],
                    "happened_at": r[1].astimezone(JST).strftime("%Y-%m-%d %H:%M:%S") if r[1] else None,
                    "target": r[2],
                    "op": r[3],
                    "batch_id": str(r[4]) if r[4] else None,
                    "actor_username": r[5],
                    "client_ip": r[6],
                    "user_agent": r[7],
                    "record": json.dumps(rec, ensure_ascii=False),  # ←文字列にして [object Object] 回避
                })

            fixed_cols = ["id","happened_at","target","op","batch_id","actor_username","client_ip","user_agent","record"]
            return {"rows": out, "columns": fixed_cols}


        rows = conn.execute(
            f"select payload from {target} order by row_no nulls last"
        ).fetchall()
        data_rows = [r[0] for r in rows]
        columns = _columns_from_schema(conn, target)

    return {"rows": data_rows, "columns": columns}


from collections import Counter
import json

def _canon_row(row: dict) -> str:
    # dict のキー順の違いを吸収するため、JSONをソートして固定文字列にする
    # ※ 値の "" と None は別物のまま（あなたの「全フィールド同じ」定義に忠実）
    return json.dumps(row, ensure_ascii=False, sort_keys=True, separators=(",", ":"))

def _diff_by_fullmatch(old_rows: list[dict], new_rows: list[dict]) -> dict:
    old_c = Counter(_canon_row(r) for r in old_rows if isinstance(r, dict))
    new_c = Counter(_canon_row(r) for r in new_rows if isinstance(r, dict))

    added = []
    deleted = []

    # 追加（新の方が多い分）
    for k, n in (new_c - old_c).items():
        row = json.loads(k)
        for _ in range(n):
            added.append(row)

    # 削除（旧の方が多い分）
    for k, n in (old_c - new_c).items():
        row = json.loads(k)
        for _ in range(n):
            deleted.append(row)

    return {
        "added_count": len(added),
        "deleted_count": len(deleted),
        "added": added,
        "deleted": deleted,
    }


@app.post("/api/save/{target}")
def save_table(target: str, request: Request, payload: Any = Body(...)):
    if target not in ALLOWED_TARGETS:
        raise HTTPException(400, f"target must be one of {ALLOWED_TARGETS}")

    _check_table_permission(request, target, write=True)

    # 受け取り：配列 or {rows:[...], columns:[...]}
    if isinstance(payload, list):
        rows = payload
        columns = []
    elif isinstance(payload, dict) and isinstance(payload.get("rows"), list):
        rows = payload["rows"]
        columns = payload.get("columns", [])
    else:
        raise HTTPException(422, "Invalid body. Expected a JSON array or {rows:[...]}")

    # columns が無い場合は rows から推定（キーの和集合）
    if not (isinstance(columns, list) and all(isinstance(c, str) for c in columns) and columns):
        seen = []
        seen_set = set()
        for r in rows:
            if isinstance(r, dict):
                for k in r.keys():
                    if k not in seen_set:
                        seen.append(k)
                        seen_set.add(k)
        columns = seen

    # ===== login_users のときだけ：平文password / password_hash を hash 化して保存 =====
    if target == LOGIN_USERS_TARGET:
        def _is_already_hashed(v) -> bool:
            s = "" if v is None else str(v).strip()
            return s.startswith("pbkdf2_sha256$")

        cols = [c for c in columns if c != "password"]
        if "password_hash" not in cols:
            cols.append("password_hash")
        columns = cols

        new_rows = []
        for row in rows:
            if not isinstance(row, dict):
                new_rows.append(row)
                continue

            row2 = dict(row)

            pw = row2.get("password")
            if pw is not None and str(pw).strip() != "":
                row2["password_hash"] = hash_password(str(pw))
            elif ("password_hash" in row2
                  and row2["password_hash"] is not None
                  and str(row2["password_hash"]).strip() != ""
                  and not _is_already_hashed(row2["password_hash"])):
                # password_hash に平文を入れても hash化して置き換える（運用ラク）
                row2["password_hash"] = hash_password(str(row2["password_hash"]))

            if "password" in row2:
                del row2["password"]

            new_rows.append(row2)

        rows = new_rows

    # ===== schedule のときだけ：「サイズ」を半角大文字で保存 =====
    if target == "schedule":
        normed = []
        for row in rows:
            if not isinstance(row, dict):
                normed.append(row)
                continue

            row2 = dict(row)

            if "サイズ" in row2:
                v = row2.get("サイズ")
                if v is None:
                    row2["サイズ"] = None
                else:
                    s = unicodedata.normalize("NFKC", str(v).strip())  # 全角→半角
                    row2["サイズ"] = s.upper()                         # 大文字化

            normed.append(row2)

        rows = normed


    actor = session_user(request) or {}
    actor_username = actor.get("username")

    with connect_db() as conn:
        ensure_tables(conn)

        # ★旧データ取得（deleteの前！）
        old = conn.execute(f"select payload from {target} order by row_no nulls last").fetchall()
        old_rows = [r[0] for r in old]

        # ★新データ（空行スキップ後）を作る
        new_rows = []
        for row in rows:
            if isinstance(row, dict) and _is_empty_row(row):   # 既存の空行スキップ
                continue
            new_rows.append(row)

        diff = _diff_by_fullmatch(old_rows, new_rows)


        conn.execute("delete from table_schema where target=%s", (target,))
        conn.execute(
            "insert into table_schema (target, columns) values (%s, %s)",
            (target, Jsonb(columns)),
        )

        conn.execute(f"delete from {target};")

        row_no = 1
        for row in rows:
            if isinstance(row, dict) and _is_empty_row(row):
                continue
            conn.execute(
                f"insert into {target} (row_no, payload) values (%s, %s)",
                (row_no, Jsonb(row)),
            )
            row_no += 1

        saved_count = row_no - 1

        client_ip = _pick_client_ip(request)
        user_agent = request.headers.get("user-agent")

        batch_id = uuid.uuid4()

        # 追加されたレコードごとに1行
        for row in (diff.get("added") or []):
            conn.execute(
                f"""
                insert into {UPDATE_LOG_TARGET}
                (target, saved_rows, columns, actor_username, client_ip, user_agent, batch_id, op, record)
                values
                (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (target, saved_count, Jsonb(columns), actor_username, client_ip, user_agent,
                batch_id, "add", Jsonb(row)),
            )

        # 削除されたレコードごとに1行
        for row in (diff.get("deleted") or []):
            conn.execute(
                f"""
                insert into {UPDATE_LOG_TARGET}
                (target, saved_rows, columns, actor_username, client_ip, user_agent, batch_id, op, record)
                values
                (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (target, saved_count, Jsonb(columns), actor_username, client_ip, user_agent,
                batch_id, "delete", Jsonb(row)),
            )


        conn.commit()

    return {"ok": True, "saved": saved_count, "columns": columns}


# ===== Excelエクスポート（ログイン必須だけ） =====

def _payload_get(payload: Any, keys: List[str]) -> str:
    if not isinstance(payload, dict):
        return ""
    for k in keys:
        if k in payload:
            v = payload.get(k)
            s = "" if v is None else str(v).strip()
            if s != "":
                return s
    return ""


def _to_int_or_none(s: str) -> Optional[int]:
    s = (s or "").strip()
    if s == "":
        return None
    try:
        s2 = s.translate(str.maketrans("０１２３４５６７８９", "0123456789"))
        return int(s2)
    except Exception:
        return None


def _copy_row_style(ws, src_row: int, dst_row: int, max_col: int) -> None:
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for c in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=c)
        dst = ws.cell(row=dst_row, column=c)
        if isinstance(dst, MergedCell):
            continue
        dst._style = copy(src._style)
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)
        dst.alignment = copy(src.alignment)


def _unmerge_intersecting(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for cr in list(ws.merged_cells.ranges):
        if not (cr.max_row < min_row or cr.min_row > max_row or cr.max_col < min_col or cr.min_col > max_col):
            ws.unmerge_cells(str(cr))


def _copy_merges_from_row(ws, src_row: int, dst_row: int, max_col: int) -> None:
    merges = list(ws.merged_cells.ranges)
    for cr in merges:
        if not (cr.min_row <= src_row <= cr.max_row):
            continue
        if cr.max_col < 1 or cr.min_col > max_col:
            continue

        height = cr.max_row - cr.min_row
        row_offset = src_row - cr.min_row

        dst_min_row = dst_row - row_offset
        dst_max_row = dst_min_row + height

        dst_min_col = max(1, cr.min_col)
        dst_max_col = min(max_col, cr.max_col)

        _unmerge_intersecting(ws, dst_min_row, dst_max_row, dst_min_col, dst_max_col)

        rng = (
            f"{get_column_letter(dst_min_col)}{dst_min_row}:"
            f"{get_column_letter(dst_max_col)}{dst_max_row}"
        )
        ws.merge_cells(rng)

def _merge_if_valid(ws, cr: CellRange):
    # 範囲が壊れてたら捨てる
    if cr.min_row < 1 or cr.min_col < 1:
        return
    if cr.max_row < cr.min_row or cr.max_col < cr.min_col:
        return
    # 1セル結合は不要（mergeしない）
    if cr.min_row == cr.max_row and cr.min_col == cr.max_col:
        return
    ws.merge_cells(
        start_row=cr.min_row, start_column=cr.min_col,
        end_row=cr.max_row, end_column=cr.max_col
    )

def _safe_delete_cols(ws, start_col: int, end_col: int) -> None:
    if start_col > end_col:
        return
    del_cnt = end_col - start_col + 1

    old_merges = [CellRange(str(r)) for r in list(ws.merged_cells.ranges)]

    # 全解除
    for r in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(r))

    # 削除
    ws.delete_cols(start_col, del_cnt)

    # 復元
    for cr in old_merges:
        # 削除範囲と重なる結合は捨てる
        if not (cr.max_col < start_col or cr.min_col > end_col):
            continue

        # 右側は左へシフト
        if cr.min_col > end_col:
            cr.min_col -= del_cnt
            cr.max_col -= del_cnt

        _merge_if_valid(ws, cr)



def _safe_delete_rows(ws, start_row: int, end_row: int) -> None:
    if start_row > end_row:
        return
    del_cnt = end_row - start_row + 1

    old_merges = [CellRange(str(r)) for r in list(ws.merged_cells.ranges)]

    for r in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(r))

    ws.delete_rows(start_row, del_cnt)

    for cr in old_merges:
        # 削除範囲と重なる結合は捨てる
        if not (cr.max_row < start_row or cr.min_row > end_row):
            continue

        # 下側は上へシフト
        if cr.min_row > end_row:
            cr.min_row -= del_cnt
            cr.max_row -= del_cnt

        _merge_if_valid(ws, cr)

from openpyxl.styles.colors import COLOR_INDEX

def _rgb_hex_from_openpyxl_color(color) -> Optional[str]:
    if not color:
        return None

    try:
        ctype = getattr(color, "type", None)
    except Exception:
        return None

    # 1) rgb (例: "FF112233" or "112233")
    if ctype == "rgb":
        rgb = getattr(color, "rgb", None)
        if isinstance(rgb, str) and re.fullmatch(r"[0-9A-Fa-f]{8}|[0-9A-Fa-f]{6}", rgb):
            return "#" + rgb[-6:]
        return None

    # 2) indexed (Excelのパレット)
    if ctype == "indexed":
        idx = getattr(color, "indexed", None)
        if isinstance(idx, int) and 0 <= idx < len(COLOR_INDEX):
            return "#" + COLOR_INDEX[idx][-6:]
        # 64などは "automatic" 扱いが多いので不明 → None
        return None

    # 3) theme / auto など（完全再現は重いのでここでは諦め）
    return None


def _border_side_to_css(side) -> Optional[str]:
    if not side:
        return None
    style = getattr(side, "style", None)
    if not style:
        return None

    width = "1px"
    if style in ("medium", "thick"):
        width = "2px"

    line = "solid"
    if style == "dotted":
        line = "dotted"
    elif style in ("dashed", "dashDot", "dashDotDot"):
        line = "dashed"

    col = _rgb_hex_from_openpyxl_color(getattr(side, "color", None)) or "#000000"
    return f"{width} {line} {col}"






# ===== Excel(xlsx) を「Tabulatorプレビュー用JSON」に変換 =====
def _xlsx_to_preview(content: bytes, data_only: bool = True) -> dict:
    """
    content: xlsx bytes
    data_only: Trueならセル値（Excel保存済みの計算結果）、Falseなら数式文字列も返す
    """
    wb = load_workbook(io.BytesIO(content), data_only=data_only)

    # 最初の visible シート
    ws = None
    for s in wb.worksheets:
        if getattr(s, "sheet_state", "visible") == "visible":
            ws = s
            break
    if ws is None:
        ws = wb.active

    MAX_COL = column_index_from_string("BY")
    MAX_ROW = 1200

    max_row = min(ws.max_row or 1, MAX_ROW)
    max_col = min(ws.max_column or 1, MAX_COL)

    # merged_map[(r,c)] = (min_row, min_col, max_row, max_col)
    merged_map: dict[tuple[int, int], tuple[int, int, int, int]] = {}
    for rg in list(ws.merged_cells.ranges):
        mr1, mc1, mr2, mc2 = rg.min_row, rg.min_col, rg.max_row, rg.max_col
        for rr in range(mr1, mr2 + 1):
            for cc in range(mc1, mc2 + 1):
                merged_map[(rr, cc)] = (mr1, mc1, mr2, mc2)

    # ===== DataValidation(list) を抽出して返す =====
    def _split_explicit_list(s: str) -> list[str]:
        s = (s or "").strip()
        if len(s) >= 2 and s[0] == '"' and s[-1] == '"':
            s = s[1:-1]
        sep = "," if "," in s else (";" if ";" in s else ",")
        return [x.strip() for x in s.split(sep) if x.strip()]

    def _values_from_range(ws_any, a1_range: str) -> list[str]:
        # "$A$1:$A$10" などを想定
        a1_range = (a1_range or "").replace("$", "").strip()
        try:
            min_col, min_row, max_col, max_row2 = range_boundaries(a1_range)
        except Exception:
            return []
        out = []
        seen = set()
        for rr in range(min_row, max_row2 + 1):
            for cc in range(min_col, max_col + 1):
                v = ws_any.cell(row=rr, column=cc).value
                if v is None:
                    continue
                s = str(v).strip()
                if not s:
                    continue
                if s in seen:
                    continue
                seen.add(s)
                out.append(s)
        return out

    def _resolve_list_values(formula1) -> list[str]:
        f = "" if formula1 is None else str(formula1).strip()
        if not f:
            return []
        if f.startswith("="):
            f = f[1:].strip()

        # 1) "A,B,C" の直書き
        if len(f) >= 2 and f[0] == '"' and f[-1] == '"':
            return _split_explicit_list(f)

        # 2) シート参照 or 同一シート範囲（A1:A10 など）
        #    例: Sheet1!$A$1:$A$10 / 'Sheet 1'!$A$1:$A$10 / $A$1:$A$10
        if ":" in f:
            sheet_name = ws.title
            ref = f
            if "!" in f:
                sp, ref = f.split("!", 1)
                sheet_name = sp.strip().strip("'")
            try:
                ws2 = wb[sheet_name]
            except Exception:
                ws2 = ws
            vals = _values_from_range(ws2, ref)
            if vals:
                return vals

        # 3) 定義名（Named Range）
        name = f.strip()
        try:
            dn = wb.defined_names.get(name)
        except Exception:
            dn = None
        if dn:
            vals = []
            seen = set()
            for title, ref in list(dn.destinations):
                ws2 = wb[title] if title in wb.sheetnames else ws
                for s in _values_from_range(ws2, ref):
                    if s not in seen:
                        seen.add(s)
                        vals.append(s)
            if vals:
                return vals

        # INDIRECT 等の複雑系は今回は未対応（空で返す）
        return []

    # ===== このアプリが出力したExcel(_dv_lists)から候補リストを復元 =====
    def _extract_dv_lists_from_hidden_sheet():
        if "_dv_lists" not in wb.sheetnames:
            return {"busu_map": {}, "pattern_map": {}}

        ws_list = wb["_dv_lists"]

        # 1行目ヘッダ: range_name -> 列番号
        header_to_col = {}
        for cc in range(1, (ws_list.max_column or 1) + 1):
            v = ws_list.cell(row=1, column=cc).value
            if v is None:
                continue
            s = str(v).strip()
            if s:
                header_to_col[s] = cc

        def read_list(range_name: str) -> list[str]:
            col = header_to_col.get(range_name)
            if not col:
                return []
            out = []
            for rr in range(2, (ws_list.max_row or 1) + 1):
                v = ws_list.cell(row=rr, column=col).value
                if v is None:
                    break
                s = str(v).strip()
                if s == "":
                    break
                out.append(s)
            return out

        busu_map = {}
        pattern_map = {}

        # A:B = 店番 -> BUSU_xxx
        for rr in range(2, (ws_list.max_row or 1) + 1):
            sc = ws_list.cell(row=rr, column=1).value
            rn = ws_list.cell(row=rr, column=2).value
            if sc is None and rn is None:
                continue
            scs = "" if sc is None else str(sc).strip()
            rns = "" if rn is None else str(rn).strip()
            if scs and rns:
                busu_map[scs] = read_list(rns)

        # C:D = 店番|部数 -> PAT_xxx
        for rr in range(2, (ws_list.max_row or 1) + 1):
            key = ws_list.cell(row=rr, column=3).value
            rn = ws_list.cell(row=rr, column=4).value
            if key is None and rn is None:
                continue
            ks = "" if key is None else str(key).strip()
            rns = "" if rn is None else str(rn).strip()
            if ks and rns:
                pattern_map[ks] = read_list(rns)

        return {"busu_map": busu_map, "pattern_map": pattern_map}


    validations_by_col: dict[str, list[dict]] = defaultdict(list)

    try:
        dvs = list(getattr(ws.data_validations, "dataValidation", []) or [])
    except Exception:
        dvs = []

    dv_maps = _extract_dv_lists_from_hidden_sheet()
    busu_cols = set()
    pattern_cols = set()

    for dv in dvs:
        if getattr(dv, "type", None) != "list":
            continue

        f1 = "" if getattr(dv, "formula1", None) is None else str(dv.formula1)
        f1s = f1.replace(" ", "")

        is_busu_formula = ("_dv_lists" in f1) and ("$A:$B" in f1)
        is_pat_formula  = ("_dv_lists" in f1) and ("$C:$D" in f1)

        allow_blank = bool(getattr(dv, "allow_blank", False))

        sqref = getattr(dv, "sqref", None)
        if not sqref:
            continue

        # ★式ベースは values を解決できないので列だけ記録（候補は dv_maps からJSで組み立てる）
        if is_busu_formula or is_pat_formula:
            for cr in list(getattr(sqref, "ranges", []) or []):
                r1, r2 = cr.min_row, cr.max_row
                c1, c2 = cr.min_col, cr.max_col
                r1 = max(1, min(r1, max_row)); r2 = max(1, min(r2, max_row))
                c1 = max(1, min(c1, max_col)); c2 = max(1, min(c2, max_col))
                if r1 > r2 or c1 > c2:
                    continue
                for cc in range(c1, c2 + 1):
                    col_letter = get_column_letter(cc)
                    if is_busu_formula:
                        busu_cols.add(col_letter)
                    if is_pat_formula:
                        pattern_cols.add(col_letter)
            continue

        # ★直書き/範囲/定義名の “静的リスト” は従来どおり values を返す
        values = _resolve_list_values(getattr(dv, "formula1", None))
        if not values:
            continue

        for cr in list(getattr(sqref, "ranges", []) or []):
            r1, r2 = cr.min_row, cr.max_row
            c1, c2 = cr.min_col, cr.max_col
            r1 = max(1, min(r1, max_row)); r2 = max(1, min(r2, max_row))
            c1 = max(1, min(c1, max_col)); c2 = max(1, min(c2, max_col))
            if r1 > r2 or c1 > c2:
                continue
            for cc in range(c1, c2 + 1):
                col_letter = get_column_letter(cc)
                validations_by_col[col_letter].append({
                    "r1": r1, "r2": r2,
                    "values": values,
                    "allow_blank": allow_blank,
                })

    print("[import_xlsx] dvs=", len(dvs),
          "static_cols=", len(validations_by_col),
          "busu_cols=", len(busu_cols),
          "pattern_cols=", len(pattern_cols),
          "busu_map_keys=", len(dv_maps.get("busu_map", {})),
          "pattern_map_keys=", len(dv_maps.get("pattern_map", {})))


    def _as_iso(v):
        # Excelの日付セルは openpyxl だと datetime になることが多く、
        # そのままだと "2025-01-01T00:00:00" のように時刻まで表示される。
        # ここでは日付だけ (YYYY-MM-DD) を返して表示をスッキリさせる。
        if isinstance(v, _dt.datetime):
            return v.date().isoformat()
        if isinstance(v, _dt.date):
            return v.isoformat()
        return v


    columns = [get_column_letter(c) for c in range(1, max_col + 1)]
    rows = []
    styles = {}  # { "row": { "A": {css...}, ... } }

    for r in range(1, max_row + 1):
        obj = {"__r": r}
        row_style = {}

        for c in range(1, max_col + 1):
            key = get_column_letter(c)

            in_merge = (r, c) in merged_map

            if in_merge:
                mr1, mc1, mr2, mc2 = merged_map[(r, c)]
                tl_cell = ws.cell(row=mr1, column=mc1)   # 値/塗りは左上
                edge_cell = ws.cell(row=r, column=c)     # 罫線は外周セル側で判断

                # 値：左上だけ表示（他は空）
                if (r, c) == (mr1, mc1):
                    tlv = tl_cell.value
                    # 数式はプレビューでは表示しない（重い計算も回避）
                    if getattr(tl_cell, "data_type", None) == "f" or (isinstance(tlv, str) and tlv.startswith("=")):
                        tlv = None
                    obj[key] = _as_iso(tlv)
                else:
                    obj[key] = None

                css = {}

                # 塗り：結合範囲全体に左上の塗りを適用
                fill = getattr(tl_cell, "fill", None)
                if fill and getattr(fill, "patternType", None):
                    bg = _rgb_hex_from_openpyxl_color(getattr(fill, "fgColor", None))
                    if bg:
                        css["backgroundColor"] = bg

                # 罫線：結合範囲の「外周」だけ出す（内側の罫線は捨てる）
                bd = getattr(edge_cell, "border", None)
                if bd:
                    if r == mr1:
                        top = _border_side_to_css(getattr(bd, "top", None))
                        if top:
                            css["borderTop"] = top
                    if r == mr2:
                        bottom = _border_side_to_css(getattr(bd, "bottom", None))
                        if bottom:
                            css["borderBottom"] = bottom
                    if c == mc1:
                        left = _border_side_to_css(getattr(bd, "left", None))
                        if left:
                            css["borderLeft"] = left
                    if c == mc2:
                        right = _border_side_to_css(getattr(bd, "right", None))
                        if right:
                            css["borderRight"] = right

                if css:
                    row_style[key] = css

            else:
                cell = ws.cell(row=r, column=c)
                v = cell.value
                # 数式はプレビューでは表示しない（重い計算も回避）
                if getattr(cell, "data_type", None) == "f" or (isinstance(v, str) and v.startswith("=")):
                    v = None
                obj[key] = _as_iso(v)

                css = {}

                # 背景色
                fill = getattr(cell, "fill", None)
                if fill and getattr(fill, "patternType", None):
                    bg = _rgb_hex_from_openpyxl_color(getattr(fill, "fgColor", None))
                    if bg:
                        css["backgroundColor"] = bg

                # 罫線（通常セルはそのまま）
                bd = getattr(cell, "border", None)
                if bd:
                    top = _border_side_to_css(getattr(bd, "top", None))
                    right = _border_side_to_css(getattr(bd, "right", None))
                    bottom = _border_side_to_css(getattr(bd, "bottom", None))
                    left = _border_side_to_css(getattr(bd, "left", None))
                    if top:
                        css["borderTop"] = top
                    if right:
                        css["borderRight"] = right
                    if bottom:
                        css["borderBottom"] = bottom
                    if left:
                        css["borderLeft"] = left

                if css:
                    row_style[key] = css

        if row_style:
            styles[str(r)] = row_style

        rows.append(obj)

    return {
        "sheet": ws.title,
        "columns": columns,
        "rows": rows,
        "styles": styles,
        "max_row": max_row,
        "max_col": max_col,
        "validations_by_col": validations_by_col,
        "dv_maps": {
            "busu_map": dv_maps.get("busu_map", {}),
            "pattern_map": dv_maps.get("pattern_map", {}),
            "busu_cols": sorted(list(busu_cols)),
            "pattern_cols": sorted(list(pattern_cols)),
        },
    }



@app.post("/api/import-xlsx")
def import_xlsx(request: Request, file: UploadFile = File(...)):
    # 読み取りできる人ならOK
    require_role(request, "viewer")


    name = (file.filename or "").lower()
    if not name.endswith(".xlsx"):
        raise HTTPException(400, "xlsxファイルを選んでください")

    content = file.file.read()
    if not content:
        raise HTTPException(400, "ファイルが空です")

    if len(content) > 15 * 1024 * 1024:
        raise HTTPException(400, "ファイルが大きすぎます（15MBまで）")

    # アップロードされたExcelをそのままプレビュー用JSONへ
    pid, meta = _preview_cache_put(content)
    out = _xlsx_to_preview(content, data_only=True)
    # 初期表示も計算結果を埋める（数式文字列ではなく値を見せる）
    if _HAS_FORMULAS and PREVIEW_CALC_ON_LOAD:
        try:
            _ups = _evaluate_formula_cells(content, out.get("sheet") or meta.get("sheet") or "", meta.get("formula_cells") or [])
            for u in _ups:
                r = int(u.get("r") or 0)
                c = str(u.get("c") or "").upper()
                if 1 <= r <= len(out.get("rows") or []) and c:
                    out["rows"][r-1][c] = u.get("v")
        except Exception:
            pass
    out["preview_id"] = pid
    out["formula_cells"] = [{"r": r, "c": c} for (r, c) in (meta.get("formula_cells") or [])]
    out["calc_supported"] = bool(_HAS_FORMULAS and ENABLE_PREVIEW_RECALC)
    return out
# ===== Excelダウンロードと同じ加工を、プレビュー用JSONにも使い回す =====
def _export_xlsx_bytes() -> bytes:
    """export_xlsx と同じ加工済み list_format.xlsx を bytes で返す"""
    if not XLSX_PATH.is_file():
        raise HTTPException(404, f"list_format.xlsx not found next to app.py: {XLSX_PATH}")

    with connect_db() as conn:
        ensure_tables(conn)
        db_rows = conn.execute(
            "select payload from shop_master order by row_no nulls last"
        ).fetchall()
        schedule_rows = conn.execute(
            "select payload from schedule order by row_no nulls last"
        ).fetchall()
        syoken_rows = conn.execute(
            "select payload from syoken order by row_no nulls last"
        ).fetchall()

    SHOP_CODE_KEYS = ["店番", "店番フィールド", "shop_code", "shop_cd", "店舗コード", "店コード", "店舗番号"]
    SHOP_NAME_KEYS = ["店名", "店名フィールド", "店舗名", "shop_name", "店舗"]
    B_ALL_KEYS = ["B全", "B全フィールド", "B_all", "B_ALL", "B1", "B1フィールド"]
    B2_KEYS = ["B2", "B2フィールド"]
    B3_KEYS = ["B3", "B3フィールド"]
    B4_KEYS = ["B4", "B4フィールド"]
    GROUP_NO_KEYS = ["グループ番", "グループ番フィールド", "group_no", "group_num", "グループ番号"]
    GROUP_NAME_KEYS = ["グループ", "グループフィールド", "group", "group_name"]

    SCHEDULE_TITLE_KEYS = ["タイトル", "タイトルフィールド", "title", "タイトル名"]
    SCHEDULE_FOLD_KEYS  = ["折込日", "折込日フィールド", "fold_date", "fold", "折込"]
    SCHEDULE_START_KEYS = ["開始日", "開始日フィールド", "start_date", "start", "開始"]
    SCHEDULE_END_KEYS   = ["終了日", "終了日フィールド", "end_date", "end", "終了"]
    SCHEDULE_SIZE_KEYS  = ["サイズ", "サイズフィールド", "size"]

    SYOKEN_SHOP_CODE_KEYS = ["店番", "店番フィールド", "shop_code", "shop_cd", "店舗コード", "店コード", "店舗番号"]
    SYOKEN_BUSU_KEYS = ["部数", "部数フィールド", "部数(折込)", "busu", "copies", "qty", "数量"]
    SYOKEN_PATTERN_KEYS = ["パターン", "パターンフィールド", "pattern", "pattern_name", "パターン名"]

    def _payload_get_any(payload: Any, keys: List[str]):
        if not isinstance(payload, dict):
            return None
        for k in keys:
            if k in payload:
                v = payload.get(k)
                if v is None:
                    continue
                if isinstance(v, str) and v.strip() == "":
                    continue
                return v
        return None

    def _to_excel_date_value(v):
        import datetime as _dt
        import re as _re
        if v is None:
            return None
        if isinstance(v, _dt.datetime):
            return v.date()
        if isinstance(v, _dt.date):
            return v
        if isinstance(v, (int, float)):
            return v

        s = str(v).strip()
        if s == "":
            return None

        s2 = s.translate(str.maketrans("０１２３４５６７８９", "0123456789"))

        m = _re.fullmatch(r"(\d{4})年(\d{1,2})月(\d{1,2})日", s2)
        if m:
            try:
                return _dt.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            except Exception:
                return s

        for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d", "%Y%m%d"):
            try:
                return _dt.datetime.strptime(s2, fmt).date()
            except Exception:
                pass

        n = _as_number_or_text(s2)
        if isinstance(n, (int, float)):
            return n
        return s

    busu_map = defaultdict(list)              # shop_code -> [busu...]
    pattern_map = defaultdict(list)           # (shop_code, busu_key) -> [pattern...]

    def _norm_shop_code(x: str) -> str:
        s = ("" if x is None else str(x)).strip()
        s = s.translate(str.maketrans("０１２３４５６７８９", "0123456789"))
        return s

    def _norm_busu_key(x: str) -> str:
        s = ("" if x is None else str(x)).strip()
        s = s.translate(str.maketrans("０１２３４５６７８９，", "0123456789,"))  # 全角→半角
        s = s.replace(",", "")  # 1,000 → 1000
        n = _as_number_or_text(s)
        # 数値なら "1000" みたいに文字列化（Excel側で H&"" した時と合わせる）
        if isinstance(n, (int, float)):
            # 余計な .0 を避けたいなら int 判定を強めてもOK
            if isinstance(n, float) and n.is_integer():
                n = int(n)
            return str(n)
        return s

    for (payload,) in syoken_rows:
        payload = payload or {}

        sc_raw = _payload_get(payload, SYOKEN_SHOP_CODE_KEYS)
        bs_raw = _payload_get_any(payload, SYOKEN_BUSU_KEYS)
        pt_raw = _payload_get_any(payload, SYOKEN_PATTERN_KEYS)

        sc = _norm_shop_code(sc_raw)
        if sc == "":
            continue
        if bs_raw is None:
            continue

        bs_key = _norm_busu_key(bs_raw)
        if bs_key == "":
            continue

        # H列用：店番ごとの部数候補
        if bs_key not in busu_map[sc]:
            busu_map[sc].append(bs_key)

        # I列用： (店番, 部数) ごとのパターン候補
        if pt_raw is not None:
            pt = str(pt_raw).strip()
            if pt != "" and pt not in pattern_map[(sc, bs_key)]:
                pattern_map[(sc, bs_key)].append(pt)


    def _sort_key(v: str):
        n = _as_number_or_text(v)
        if isinstance(n, (int, float)):
            return (0, float(n))
        return (1, v)

    for sc in list(busu_map.keys()):
        busu_map[sc] = sorted(busu_map[sc], key=_sort_key)

    # ===== 印刷単価（shop_master の「グループ=印刷単価」レコード）を探す =====
    PRINT_GROUP_NAME = "印刷単価"
    print_price_payload = None
    for (payload,) in db_rows:
        payload = payload or {}
        gname = _payload_get(payload, GROUP_NAME_KEYS)  # "グループ" 系
        if (gname or "").strip() == PRINT_GROUP_NAME:
            print_price_payload = payload
            break

    # サイズ -> 単価
    price_map = {}
    if isinstance(print_price_payload, dict):
        # ここはあなたのフィールド定義に合わせて増減してOK
        price_map = {
            "B全": _as_number_or_text(_payload_get(print_price_payload, B_ALL_KEYS)),
            "B2":  _as_number_or_text(_payload_get(print_price_payload, B2_KEYS)),
            "B3":  _as_number_or_text(_payload_get(print_price_payload, B3_KEYS)),
            "B4":  _as_number_or_text(_payload_get(print_price_payload, B4_KEYS)),
        }
        # 空は落とす
        price_map = {k: v for k, v in price_map.items() if v is not None and str(v).strip() != ""}


    items = []
    for (payload,) in db_rows:
        payload = payload or {}
        shop_code_v = _payload_get(payload, SHOP_CODE_KEYS)
        shop_name_v = _payload_get(payload, SHOP_NAME_KEYS)
        b_all_v = _payload_get(payload, B_ALL_KEYS)
        b2_v = _payload_get(payload, B2_KEYS)
        b3_v = _payload_get(payload, B3_KEYS)
        b4_v = _payload_get(payload, B4_KEYS)
        group_no_v = _payload_get(payload, GROUP_NO_KEYS)
        group_name_v = _payload_get(payload, GROUP_NAME_KEYS)

        group_no_int = _to_int_or_none(group_no_v)
        items.append({
            "shop_code": shop_code_v,
            "shop_name": shop_name_v,
            "b_all": b_all_v,
            "b2": b2_v,
            "b3": b3_v,
            "b4": b4_v,
            "group_no": group_no_v,
            "group_no_int": group_no_int,
            "group_name": group_name_v,
        })

    filtered = []
    for it in items:
        if it.get("group_no_int") == 0:
            continue
        if (it.get("group_no") or "").strip() in ("0", "０"):
            continue
        filtered.append(it)

    out_rows: List[Dict[str, str]] = []
    current_marker = None
    current_label = ""
    group_shop_rows: List[Dict[str, str]] = []

    def flush_group():
        nonlocal group_shop_rows, current_label
        if not group_shop_rows:
            return
        out_rows.extend(group_shop_rows)
        if (current_label or "").strip() != "":
            out_rows.append({"A": "G", "B": "", "C": current_label, "D": "", "E": "", "F": "", "G": ""})
        group_shop_rows = []

    for it in filtered:
        marker = it["group_no"] or it["group_name"] or ""
        label = it["group_name"] or it["group_no"] or ""

        if current_marker is None:
            current_marker = marker
            current_label = label
        elif marker != current_marker:
            flush_group()
            current_marker = marker
            current_label = label

        group_shop_rows.append({
            "A": "D",
            "B": it["shop_code"],
            "C": it["shop_name"],
            "D": it["b_all"],
            "E": it["b2"],
            "F": it["b3"],
            "G": it["b4"],
        })

    flush_group()

    def _set_value_safe(ws, row: int, col: int, value):
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            if value is None or value == "":
                return
            for cr in ws.merged_cells.ranges:
                if cr.min_row <= row <= cr.max_row and cr.min_col <= col <= cr.max_col:
                    ws.cell(row=cr.min_row, column=cr.min_col).value = value
                    return
            return
        cell.value = value

    wb = load_workbook(str(XLSX_PATH))
    ws = wb.active

    base_col = column_index_from_string("H")
    step = 4
    row_start = 5
    row_title = 6
    row_size = 7
    max_col = column_index_from_string("BY")

    sch_items = []
    for (payload,) in schedule_rows:
        payload = payload or {}
        title_v = _payload_get_any(payload, SCHEDULE_TITLE_KEYS)
        fold_v  = _payload_get_any(payload, SCHEDULE_FOLD_KEYS)
        start_v = _payload_get_any(payload, SCHEDULE_START_KEYS)
        end_v   = _payload_get_any(payload, SCHEDULE_END_KEYS)
        size_v  = _payload_get_any(payload, SCHEDULE_SIZE_KEYS)

        if (str(title_v or "").strip() == "" and
            str(fold_v or "").strip() == "" and
            str(start_v or "").strip() == "" and
            str(end_v or "").strip() == "" and
            str(size_v or "").strip() == ""):
            continue

        sch_items.append({
            "title": "" if title_v is None else title_v,
            "fold": fold_v,
            "start": start_v,
            "end": end_v,
            "size": "" if size_v is None else size_v,
        })

    def _to_date_only(v):
        dv = _to_excel_date_value(v)
        import datetime as _dt
        if isinstance(dv, _dt.datetime):
            return dv.date()
        if isinstance(dv, _dt.date):
            return dv
        return None

    def _fold_label(fold_v, start_v) -> str:
        fd = _to_date_only(fold_v)
        sd = _to_date_only(start_v)
        if not fd or not sd:
            return ""
        try:
            d = (sd - fd).days
        except Exception:
            return ""
        if d == 0:
            return ""
        if d == 1:
            return "前日折込"
        if d == 2:
            return "前々日折込"
        if d > 2:
            return f"{d}日前折込"
        return f"{abs(d)}日後折込"

    for i, sch in enumerate(sch_items):
        col_h = base_col + step * i
        col_k = col_h + 3
        if col_k > max_col:
            break

        _set_value_safe(ws, row_title, col_h, "" if sch["title"] is None else str(sch["title"]).strip())

        # H5: 折込日の表示（開始日との差分）
        _set_value_safe(ws, row_start, col_h, _fold_label(sch.get("fold"), sch.get("start")))

        # I5: 開始日 / K5: 終了日（J5 は区切り）
        _set_value_safe(ws, row_start, col_h + 1, _to_excel_date_value(sch["start"]))
        _set_value_safe(ws, row_start, col_h + 2, "～")
        _set_value_safe(ws, row_start, col_h + 3, _to_excel_date_value(sch["end"]))

        _set_value_safe(ws, row_size, col_h, "" if sch["size"] is None else str(sch["size"]).strip())

        # ===== サイズ入力(H7/L7/...) → 単価(K8/O8/...) を自動算出 =====
        PRICE_ROW = 8
        price_col = col_h + 3  # Hブロック→K, Lブロック→O...

        size_cell = f"{get_column_letter(col_h)}{row_size}"  # H7 / L7 ...
        # サイズが空なら空、あれば印刷単価表から引く
        formula = f'=IF({size_cell}="","",IFERROR(VLOOKUP(UPPER({size_cell}),PRINT_PRICE,2,FALSE),""))'
        _set_value_safe(ws, PRICE_ROW, price_col, formula)


    # ===== 何ブロック使ったか（=日付を書いた数）を確定 =====
    max_blocks = ((max_col - base_col) // step) + 1
    used_blocks = min(len(sch_items), max_blocks)

    if used_blocks <= 0:
        # 1個も日付が無い場合：列削除しない
        next_block_start = None
    else:
        last_block_start = base_col + step * (used_blocks - 1)   # H/L/P...
        last_end_date_col = last_block_start + 3                 # K/O/S...（終了日）
        next_block_start = last_block_start + 4                 # 次ブロック先頭（= last_block_start + 4）

    # 最後に使ったブロックの「ブロック最終列」（HブロックならK、LブロックならO…）
    last_block_endcol = None if used_blocks <= 0 else (last_block_start + 3)


    START_ROW = 10
    GROUP_STYLE_ROW = 100
    GROUP_STYLE_MAX_COL = column_index_from_string("BY")
    shop_runs = []          # [(start_row, end_row), ...] 店舗行だけの連続区間
    _run_start = None       # 現在の店舗行区間の開始行



    for i, rr in enumerate(out_rows):
        r = START_ROW + i

        # --- 店舗行(D)の連続区間を収集し、グループ行(G)で確定させる ---
        if rr.get("A") == "G":
            # 直前までの店舗区間を閉じる
            if _run_start is not None:
                shop_runs.append((_run_start, r - 1))
                _run_start = None
        else:
            # 店舗行開始
            if _run_start is None:
                _run_start = r


        if rr.get("A") == "G":
            _unmerge_intersecting(ws, r, r, 1, GROUP_STYLE_MAX_COL)
            _copy_row_style(ws, GROUP_STYLE_ROW, r, max_col=GROUP_STYLE_MAX_COL)
            _copy_merges_from_row(ws, GROUP_STYLE_ROW, r, max_col=GROUP_STYLE_MAX_COL)

            _set_value_safe(ws, r, 1, "G")
            _set_value_safe(ws, r, 3, rr.get("C", ""))

            # ★グループ行：HJK / LNO / PRS ... の各列に、直上の店舗行の列合計を入れる（BS列まで）
            if shop_runs:
                srow, erow = shop_runs[-1]
                if srow <= erow:
                    start_block = column_index_from_string("H")
                    end_col = min(column_index_from_string("BS"), last_block_endcol or column_index_from_string("K"))

                    for base in range(start_block, end_col + 1, 4):  # H, L, P, ...
                        for off in (0, 2, 3):  # HJK / LNO / PRS の「I/M/Q は除外」
                            col = base + off
                            if col > end_col:
                                continue
                            letter = get_column_letter(col)
                            _set_value_safe(ws, r, col, f"=SUM({letter}{srow}:{letter}{erow})")


        else:
            vals = [
                rr["A"],
                rr["B"],
                rr["C"],
                _as_number_or_text(rr["D"]),
                _as_number_or_text(rr["E"]),
                _as_number_or_text(rr["F"]),
                _as_number_or_text(rr["G"]),
            ]
            for c, v in enumerate(vals, start=1):
                _set_value_safe(ws, r, c, v)

    delete_to = column_index_from_string("BS")  # ← 分かりやすくBSを直書きでOK

    if next_block_start is not None and next_block_start <= delete_to:
        _safe_delete_cols(ws, next_block_start, delete_to)


    # ===== 単価表シート（hidden）を作る =====
    PRICE_SHEET = "_prices"
    if PRICE_SHEET in wb.sheetnames:
        wb.remove(wb[PRICE_SHEET])
    ws_price = wb.create_sheet(PRICE_SHEET)
    ws_price.sheet_state = "hidden"

    ws_price["A1"].value = "size"
    ws_price["B1"].value = "unit_price"

    r = 2
    for sz in ("B全", "B2", "B3", "B4"):
        if sz in price_map:
            ws_price.cell(row=r, column=1).value = sz
            ws_price.cell(row=r, column=2).value = price_map[sz]
            r += 1

    # 名前定義（VLOOKUP用）
    # データが無い場合でも壊れないように EMPTY を使う
    if r == 2:
        # データ無し
        ws_price["A2"].value = ""
        ws_price["B2"].value = ""
        r = 3

    # 例: '_prices'!$A$2:$B$5
    wb.defined_names.add(
        DefinedName("PRINT_PRICE", attr_text=f"'{PRICE_SHEET}'!$A$2:$B${r-1}")
    )


    LIST_SHEET = "_dv_lists"
    if LIST_SHEET in wb.sheetnames:
        wb.remove(wb[LIST_SHEET])

    ws_list = wb.create_sheet(LIST_SHEET)
    ws_list.sheet_state = "hidden"

    # ===== 以前の定義名が残っているとExcelが変な方を掴む事があるので掃除 =====
    for nm in list(wb.defined_names.keys()):
        if nm == "EMPTY" or nm.startswith("BUSU_") or nm.startswith("PAT_"):
            del wb.defined_names[nm]

    # ===== 店番/パターンが見つからない場合の逃げ（絶対に上書きされないセルを使う）=====
    # Excel最終列: XFD
    ws_list["XFD1"].value = ""
    wb.defined_names.add(DefinedName("EMPTY", attr_text=f"'{LIST_SHEET}'!$XFD$1:$XFD$1"))


    ws_list["A1"].value = "shop_code"
    ws_list["B1"].value = "range_name"

    ws_list["C1"].value = "shop_code|busu"
    ws_list["D1"].value = "pattern_range_name"



    # 店番ごとに 1 列ずつ値リストを作る（D列以降）
    col_ptr = 5  # E  ← Dはマッピング用に空ける
    row_ptr = 2  # マッピング表は2行目から

    for shop_code in sorted(busu_map.keys(), key=lambda x: str(x)):
        values = busu_map.get(shop_code) or []
        if not values:
            continue

        safe = str(shop_code).strip()
        safe = safe.translate(str.maketrans("０１２３４５６７８９", "0123456789"))
        safe = re.sub(r"[^A-Za-z0-9_]", "_", safe)
        if not safe or safe[0].isdigit():
            safe = "S_" + safe
        range_name = f"BUSU_{safe}"

        # マッピング表（A列=店番、B列=range_name）
        ws_list.cell(row=row_ptr, column=1).value = str(shop_code).strip()
        ws_list.cell(row=row_ptr, column=2).value = range_name
        ws_list.cell(row=row_ptr, column=1).number_format = "@"  # 店番を文字として保持

        # 値列（D列以降）
        ws_list.cell(row=1, column=col_ptr).value = range_name
        for i2, v in enumerate(values, start=2):
            ws_list.cell(row=i2, column=col_ptr).value = _as_number_or_text(v)

        last_row = 1 + len(values)
        col_letter = get_column_letter(col_ptr)
        ref = f"'{LIST_SHEET}'!${col_letter}$2:${col_letter}${last_row}"
        wb.defined_names.add(DefinedName(range_name, attr_text=ref))

        col_ptr += 1
        row_ptr += 1

    # ===== (店番|部数) ごとに パターン候補リストを作る =====
    row_ptr2 = 2  # E:F の対応表は2行目から

    def _safe_name(s: str) -> str:
        s = ("" if s is None else str(s)).strip()
        s = s.translate(str.maketrans("０１２３４５６７８９", "0123456789"))
        s = re.sub(r"[^A-Za-z0-9_]", "_", s)
        if not s or s[0].isdigit():
            s = "X_" + s
        return s

    for (sc, bs_key) in sorted(pattern_map.keys(), key=lambda x: (str(x[0]), str(x[1]))):
        patterns = pattern_map.get((sc, bs_key)) or []
        if not patterns:
            continue

        range_name = f"PAT_{_safe_name(sc)}_{_safe_name(bs_key)}"

        # 対応表（E列=店番|部数、F列=range_name）
        key = f"{sc}|{bs_key}"
        ws_list.cell(row=row_ptr2, column=3).value = key
        ws_list.cell(row=row_ptr2, column=4).value = range_name
        ws_list.cell(row=row_ptr2, column=3).number_format = "@"

        # 値列（部数リストの続きの列 col_ptr を使用）
        ws_list.cell(row=1, column=col_ptr).value = range_name
        for i2, v in enumerate(sorted(patterns, key=lambda x: str(x)), start=2):
            ws_list.cell(row=i2, column=col_ptr).value = v

        last_row = 1 + len(patterns)
        col_letter = get_column_letter(col_ptr)
        ref = f"'{LIST_SHEET}'!${col_letter}$2:${col_letter}${last_row}"
        wb.defined_names.add(DefinedName(range_name, attr_text=ref))

        col_ptr += 1
        row_ptr2 += 1



    # ===== DataValidation（H列, L列, P列... BP列に付与）=====
    dv = DataValidation(
        type="list",
        formula1=f"=INDIRECT(IFERROR(VLOOKUP(INDEX($B:$B,ROW())&\"\",'{LIST_SHEET}'!$A:$B,2,FALSE),\"EMPTY\"))",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="入力エラー",
        error="リストから選択してください。",
    )

    last_out = START_ROW + max(len(out_rows) - 1, 0)

    dv_last_row = max(last_out, GROUP_STYLE_ROW)

    # ===== DataValidation（H列, L列, P列... BP列に付与）=====
    ws.add_data_validation(dv)

    start_col = column_index_from_string("H")
    end_col = column_index_from_string("BP")
    step = 4

    for col in range(start_col, end_col + 1, step):
        col_letter = get_column_letter(col)
        for srow, erow in shop_runs:
            dv.add(f"{col_letter}{srow}:{col_letter}{erow}")

    # ===== DataValidation（I列, M列, Q列... BQ列：パターン）=====
    start_col_busu = column_index_from_string("H")
    end_col_busu   = column_index_from_string("BP")
    step = 4

    for col_busu in range(start_col_busu, end_col_busu + 1, step):
        col_pat = col_busu + 1  # H->I, L->M, P->Q ...
        if col_pat > column_index_from_string("BQ"):
            break

        busu_letter = get_column_letter(col_busu)
        pat_letter  = get_column_letter(col_pat)

        dv_pat = DataValidation(
            type="list",
            formula1=(
                f"=INDIRECT("
                f"IF("
                f"IFERROR(VLOOKUP("
                f"INDEX($B:$B,ROW())&\"|\"&INDEX(${busu_letter}:${busu_letter},ROW())&\"\","
                f"'{LIST_SHEET}'!$C:$D,2,FALSE),\"\")=\"\","
                f"\"EMPTY\","
                f"VLOOKUP("
                f"INDEX($B:$B,ROW())&\"|\"&INDEX(${busu_letter}:${busu_letter},ROW())&\"\","
                f"'{LIST_SHEET}'!$C:$D,2,FALSE)"
                f")"
                f")"
            ),
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="入力エラー",
            error="リストから選択してください。",
        )

        ws.add_data_validation(dv_pat)
        for srow, erow in shop_runs:
            dv_pat.add(f"{pat_letter}{srow}:{pat_letter}{erow}")

        # （任意）未選択で候補が複数ある時だけ薄黄色
        warn_fill = PatternFill(fill_type="solid", start_color="FFFFEB9C", end_color="FFFFEB9C")
        rule = FormulaRule(
            formula=[
                f'=AND({pat_letter}{START_ROW}="",'
                f'COUNTA(INDIRECT('
                f'IF('
                f'IFERROR(VLOOKUP('
                f'INDEX($B:$B,ROW())&"|"&INDEX(${busu_letter}:${busu_letter},ROW())&"",'
                f"'{LIST_SHEET}'!$C:$D,2,FALSE),"
                f'"" )="",'
                f'"EMPTY",'
                f'VLOOKUP('
                f'INDEX($B:$B,ROW())&"|"&INDEX(${busu_letter}:${busu_letter},ROW())&"",'
                f"'{LIST_SHEET}'!$C:$D,2,FALSE)"
                f')'
                f'))>1)'
            ],
            fill=warn_fill,
        )
        ws.conditional_formatting.add(
            f"{pat_letter}{START_ROW}:{pat_letter}{dv_last_row}",
            rule,
        )



    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()



# ===== Preview live recalculation =====
def _parse_cell_input(v):
    if v is None:
        return None
    # Tabulatorからはだいたい文字列で来る
    if isinstance(v, str):
        s = v.strip()
        if s == "":
            return None
        # ISO datetime/date
        try:
            if re.fullmatch(r"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}(?:\.\d+)?", s):
                return _dt.datetime.fromisoformat(s)
            if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
                return _dt.date.fromisoformat(s)
        except Exception:
            pass
        # 数値らしければ数値に
        n = _as_number_or_text(s)
        return n
    return v


def _jsonable(v):
    if v is None:
        return None
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.isoformat()
    # xlcalculator のエラーは str 化
    try:
        import decimal
        if isinstance(v, decimal.Decimal):
            return float(v)
    except Exception:
        pass
    # そのままJSON化できない型は文字列
    if isinstance(v, (int, float, bool, str)):
        return v
    return str(v)


def _evaluate_formula_cells(xlsx_bytes: bytes, sheet_title: str, formula_cells: list[tuple[int, str]]):
    """xlsx_bytes を formulas で再計算し、数式セルの計算結果だけ返す"""
    if not (_HAS_FORMULAS and ENABLE_PREVIEW_RECALC):
        return []

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix="preview_")
    tmp_path = tmp.name
    try:
        tmp.write(xlsx_bytes)
        tmp.close()

        xl_model = formulas.ExcelModel().loads(tmp_path).finish()
        xl_model.calculate()

        # ★ここがポイント：write() の戻り値から Workbook を直接取る
        written = xl_model.write()  # {'EXCEL.XLSX': {Book: <openpyxl Workbook ...>}} :contentReference[oaicite:1]{index=1}

        wb2 = None
        # written -> 外側dict -> 内側dict -> Workbook
        for _k, inner in (written or {}).items():
            if isinstance(inner, dict) and inner:
                wb2 = next(iter(inner.values()))
                break
        if wb2 is None:
            return []

        ws2 = wb2[sheet_title] if sheet_title in wb2.sheetnames else _first_visible_sheet(wb2)

        updates = []
        for r, col in formula_cells:
            try:
                v = ws2[f"{col}{r}"].value
            except Exception:
                v = None
            updates.append({"r": r, "c": col, "v": _jsonable(v)})
        return updates

    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass



@app.post("/api/preview-recalc")
def preview_recalc(request: Request, payload: dict = Body(...)):
    # 読み取りできる人ならOK
    require_role(request, "viewer")

    if not ENABLE_PREVIEW_RECALC:
        pid = str(payload.get("preview_id") or "").strip()
        if not pid:
            raise HTTPException(400, "preview_id がありません")
        _preview_cache_get(pid)
        return {"preview_id": pid, "updates": [], "has_calc": False}

    pid = str(payload.get("preview_id") or "").strip()
    if not pid:
        raise HTTPException(400, "preview_id がありません")

    changes = payload.get("changes") or []
    if not isinstance(changes, list):
        raise HTTPException(400, "changes は配列で送ってください")

    meta = _preview_cache_get(pid)
    base_bytes = meta["bytes"]
    sheet_title = meta.get("sheet") or ""
    formula_cells = meta.get("formula_cells") or []

    # ワークブックへ反映（数式は残す）
    wb = load_workbook(io.BytesIO(base_bytes), data_only=False)
    ws = wb[sheet_title] if sheet_title in wb.sheetnames else _first_visible_sheet(wb)
    sheet_title = ws.title  # 念のため合わせる

    for ch in changes:
        try:
            r = int(ch.get("r"))
            c = str(ch.get("c") or "").strip().upper()
        except Exception:
            continue
        if r <= 0 or not c:
            continue
        v = _parse_cell_input(ch.get("v"))
        try:
            ws[f"{c}{r}"].value = v
        except Exception:
            continue

    buf = io.BytesIO()
    wb.save(buf)
    new_bytes = buf.getvalue()

    # cache更新（入力は積み重ねる）
    meta["bytes"] = new_bytes
    meta["updated_at"] = time.time()

    updates = _evaluate_formula_cells(new_bytes, sheet_title, formula_cells)
    return {"preview_id": pid, "updates": updates, "has_calc": bool(_HAS_FORMULAS and ENABLE_PREVIEW_RECALC)}

@app.get("/api/preview-export-xlsx")
def preview_export_xlsx(request: Request):
    # 読み取りできる人ならOK
    require_role(request, "viewer")
    # Excelダウンロード（export）と同じ加工結果を、そのままプレビューとして返す
    content = _export_xlsx_bytes()
    pid, meta = _preview_cache_put(content)
    # 初期表示は数式セルを「計算結果」にしたいので、まず値プレビューを返し、クライアント側で再計算APIを叩いて埋める
    out = _xlsx_to_preview(content, data_only=True)
    # 初期表示も計算結果を埋める（数式文字列ではなく値を見せる）
    if _HAS_FORMULAS and PREVIEW_CALC_ON_LOAD:
        try:
            _ups = _evaluate_formula_cells(content, out.get("sheet") or meta.get("sheet") or "", meta.get("formula_cells") or [])
            for u in _ups:
                r = int(u.get("r") or 0)
                c = str(u.get("c") or "").upper()
                if 1 <= r <= len(out.get("rows") or []) and c:
                    out["rows"][r-1][c] = u.get("v")
        except Exception:
            pass
    out["preview_id"] = pid
    out["formula_cells"] = [{"r": r, "c": c} for (r, c) in (meta.get("formula_cells") or [])]
    out["calc_supported"] = bool(_HAS_FORMULAS and ENABLE_PREVIEW_RECALC)
    return out


@app.get("/api/export-xlsx")
def export_xlsx(background_tasks: BackgroundTasks, request: Request):
    # 読み取りできる人ならOK
    require_role(request, "viewer")

    content = _export_xlsx_bytes()

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_path = tmp.name
    tmp.write(content)
    tmp.close()

    background_tasks.add_task(lambda p: os.path.exists(p) and os.remove(p), tmp_path)

    return FileResponse(
        tmp_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="list_format.xlsx",
    )



@app.post("/api/preview-import-inputs")
def preview_import_inputs(request: Request, files: List[UploadFile] = File(...)):
    """Excelファイル(複数可)から、入力があるセルだけを抽出して返す（プレビューへ反映用）"""
    require_role(request, "viewer")

    if not isinstance(files, list) or len(files) == 0:
        raise HTTPException(400, "files がありません")

    START_ROW = 10
    MAX_ROW = 1200  # 安全上限（テンプレ側の最大行より小さくてもOK）
    start_col = column_index_from_string("H")
    end_col = column_index_from_string("BS")  # 入力列を含む範囲まで（H..BS）

    changes_map: dict[tuple[int, str], Any] = {}
    errors: list[dict] = []

    for f in files:
        name = (f.filename or "")
        if not name.lower().endswith(".xlsx"):
            errors.append({"file": name, "error": "xlsx 以外は無視しました"})
            continue

        try:
            content = f.file.read()
        except Exception:
            content = b""

        if not content:
            errors.append({"file": name, "error": "ファイルが空です"})
            continue

        try:
            wb = load_workbook(io.BytesIO(content), data_only=False)
            ws = _first_visible_sheet(wb)
            max_row = min(ws.max_row or START_ROW, MAX_ROW)

            # 4列ブロック(H..): [実施枚数, 配布エリア, 店置き, 費用] のうち
            # 先頭3列だけを対象にする（オフセット 0,1,2）
            for r in range(START_ROW, max_row + 1):
                for base in range(start_col, end_col + 1, 4):
                    for off in (0, 1, 2):
                        c = base + off
                        if c > end_col:
                            continue
                        cell = ws.cell(row=r, column=c)
                        v = cell.value

                        # 数式は「入力」とみなさない
                        if getattr(cell, "data_type", None) == "f" or (isinstance(v, str) and v.startswith("=")):
                            continue

                        if v is None:
                            continue
                        if isinstance(v, str) and v.strip() == "":
                            continue

                        col_letter = get_column_letter(c)
                        changes_map[(r, col_letter)] = _jsonable(v)

        except Exception as e:
            errors.append({"file": name, "error": f"読み込み失敗: {type(e).__name__}: {e}"})
            continue

    changes = [{"r": r, "c": c, "v": v} for (r, c), v in changes_map.items()]
    changes.sort(key=lambda x: (int(x["r"]), str(x["c"])))

    return {
        "ok": True,
        "files": len(files),
        "applied_cells": len(changes),
        "changes": changes,
        "errors": errors,
    }


@app.post("/api/export-xlsx-with-preview")
def export_xlsx_with_preview(background_tasks: BackgroundTasks, request: Request, body: dict = Body(...)):
    # 読み取りできる人ならOK
    require_role(request, "viewer")

    changes = body.get("changes") or []
    if not isinstance(changes, list):
        raise HTTPException(400, "changes は配列で送ってください")

    # まず通常のエクスポート内容を作る
    content = _export_xlsx_bytes()

    # プレビューで入力された「実施枚数/配布エリア/店置き」だけ上書きして返す
    if changes:
        wb = load_workbook(io.BytesIO(content), data_only=False)
        ws = wb.active

        # H/I/J, L/M/N, ... のみ許可（費用などは触らない）
        max_c = min(ws.max_column or 1, column_index_from_string("BS"))
        base = column_index_from_string("H")
        step = 4
        input_cols = set()
        for b in range(base, max_c + 1, step):
            for off in (0, 1, 2):  # 実施枚数/配布エリア/店置き
                c = b + off
                if c <= max_c:
                    input_cols.add(get_column_letter(c))

        for ch in changes:
            try:
                r = int(ch.get("r"))
                c = str(ch.get("c") or "").strip().upper()
            except Exception:
                continue
            if r <= 0 or not c:
                continue
            if c not in input_cols:
                continue
            v = ch.get("v")
            try:
                ws[f"{c}{r}"].value = _parse_cell_input(v)
            except Exception:
                continue

        buf = io.BytesIO()
        wb.save(buf)
        content = buf.getvalue()

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_path = tmp.name
    tmp.write(content)
    tmp.close()

    background_tasks.add_task(lambda p: os.path.exists(p) and os.remove(p), tmp_path)

    return FileResponse(
        tmp_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="list_format.xlsx",
    )


if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", "8000"))
    uvicorn.run(app, host="0.0.0.0", port=port)
